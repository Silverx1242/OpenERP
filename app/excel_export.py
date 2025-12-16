import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys
from app import database

"""
Module to export database data to Excel files.
"""

def get_export_directory():
    """
    Gets the appropriate directory to save exported files.
    On macOS, when running from a .app bundle, uses the user's directory.
    """
    # If we're running from a .app bundle on macOS
    if sys.platform == 'darwin' and '.app/Contents/MacOS' in sys.executable:
        # Use user's directory or Documents
        home_dir = os.path.expanduser('~')
        documents_dir = os.path.join(home_dir, 'Documents', 'OpenERP')
        # Create directory if it doesn't exist
        os.makedirs(documents_dir, exist_ok=True)
        return documents_dir
    else:
        # For Windows/Linux or normal execution, use current directory
        return os.getcwd()

def export_to_excel(filepath=None):
    """
    Exports all database data to an Excel file.
    Creates multiple sheets: Inventory, Sales, Revenue, Costs, BOM
    """
    if filepath is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"export_erp_{timestamp}.xlsx"
        export_dir = get_export_directory()
        filepath = os.path.join(export_dir, filename)
    else:
        # If a relative path is provided, use export directory
        if not os.path.isabs(filepath):
            export_dir = get_export_directory()
            filepath = os.path.join(export_dir, os.path.basename(filepath))
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 1. Inventory Sheet
    sheet_inventory = wb.create_sheet("Inventory")
    products = database.get_all_products()
    
    headers_inv = ['ID', 'SKU', 'Name', 'Type', 'Stock', 'Min Stock', 'Unit Cost', 
                   'Supplier', 'Purchase Date', 'Exit Date', 'Total Value']
    sheet_inventory.append(headers_inv)
    
    # Header style
    header_fill = PatternFill(start_color="4285f4", end_color="4285f4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in sheet_inventory[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for p in products:
        total_value = p['stock'] * (p['cost'] or 0)
        sheet_inventory.append([
            p['id'],
            p.get('sku', '') or '',
            p['name'],
            p['product_type'],
            p['stock'],
            p['min_stock'],
            p['cost'] or 0,
            p['supplier'] or '',
            p['purchase_date'] or '',
            p['exit_date'] or '',
            total_value
        ])
    
    # Adjust column width
    for col in range(1, len(headers_inv) + 1):
        sheet_inventory.column_dimensions[get_column_letter(col)].width = 15
    
    # 2. Sales Sheet
    sheet_sales = wb.create_sheet("Sales")
    sales = database.get_all_sales()
    
    headers_sales = ['ID', 'Product', 'Quantity', 'Unit Price', 'Total', 'Date']
    sheet_sales.append(headers_sales)
    
    for cell in sheet_sales[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for s in sales:
        sheet_sales.append([
            s['id'],
            s['product_name'],
            s['quantity'],
            s['unit_price'],
            s['total_amount'],
            s['date']
        ])
    
    for col in range(1, len(headers_sales) + 1):
        sheet_sales.column_dimensions[get_column_letter(col)].width = 18
    
    # 3. Revenue Sheet
    sheet_revenue = wb.create_sheet("Revenue")
    revenue = database.get_all_revenue()
    
    headers_revenue = ['ID', 'Description', 'Amount', 'Date']
    sheet_revenue.append(headers_revenue)
    
    for cell in sheet_revenue[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for r in revenue:
        sheet_revenue.append([
            r['id'],
            r['description'],
            r['amount'],
            r['date']
        ])
    
    for col in range(1, len(headers_revenue) + 1):
        sheet_revenue.column_dimensions[get_column_letter(col)].width = 20
    
    # 4. Costs Sheet
    sheet_costs = wb.create_sheet("Costs")
    costs = database.get_all_costs()
    
    headers_costs = ['ID', 'Description', 'Amount', 'Category', 'Date']
    sheet_costs.append(headers_costs)
    
    for cell in sheet_costs[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for c in costs:
        sheet_costs.append([
            c['id'],
            c['description'],
            c['amount'],
            c.get('category', 'Others'),
            c['date']
        ])
    
    for col in range(1, len(headers_costs) + 1):
        sheet_costs.column_dimensions[get_column_letter(col)].width = 20
    
    # 5. BOM Sheet
    sheet_bom = wb.create_sheet("BOM")
    products = database.get_all_products()
    
    headers_bom = ['Parent Product', 'Child Component', 'Quantity']
    sheet_bom.append(headers_bom)
    
    for cell in sheet_bom[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for p in products:
        bom_items = database.get_bom_for_product(p['id'])
        for bom in bom_items:
            sheet_bom.append([
                p['name'],
                bom['child_product_name'],
                bom['quantity']
            ])
    
    for col in range(1, len(headers_bom) + 1):
        sheet_bom.column_dimensions[get_column_letter(col)].width = 25
    
    # Save file
    wb.save(filepath)
    return filepath

def export_sales_report(period='month', filepath=None):
    """
    Exports a sales report filtered by period to Excel.
    period can be: 'day', 'week', 'month', 'year'
    """
    if filepath is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        period_names = {'day': 'day', 'week': 'week', 'month': 'month', 'year': 'year'}
        period_name = period_names.get(period, 'month')
        filename = f"sales_report_{period_name}_{timestamp}.xlsx"
        export_dir = get_export_directory()
        filepath = os.path.join(export_dir, filename)
    else:
        # If a relative path is provided, use export directory
        if not os.path.isabs(filepath):
            export_dir = get_export_directory()
            filepath = os.path.join(export_dir, os.path.basename(filepath))
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Get sales for the period
    sales = database.get_sales_by_period(period)
    
    # Sales Report Sheet
    sheet = wb.create_sheet("Sales Report")
    
    # Title
    period_names = {
        'day': 'Current Day',
        'week': 'Current Week',
        'month': 'Current Month',
        'year': 'Current Year'
    }
    title = period_names.get(period, 'Current Month')
    sheet.merge_cells('A1:F1')
    title_cell = sheet['A1']
    title_cell.value = f"Sales Report - {title}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ['ID', 'Date', 'Product', 'Quantity', 'Unit Price', 'Total']
    sheet.append([])  # Blank line
    sheet.append(headers)
    
    # Header style
    header_fill = PatternFill(start_color="4285f4", end_color="4285f4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in sheet[3]:  # Row 3 (after title and blank line)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    total_general = 0
    for s in sales:
        date_str = s['date'] if s['date'] else ''
        if date_str:
            try:
                if isinstance(date_str, str):
                    date_obj = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                    date_str = date_obj.strftime('%Y-%m-%d %H:%M:%S')
            except:
                pass
        
        sheet.append([
            s['id'],
            date_str,
            s['product_name'],
            s['quantity'],
            s['unit_price'],
            s['total_amount']
        ])
        total_general += s['total_amount']
    
    # Totals
    sheet.append([])
    sheet.append(['', '', '', '', 'TOTAL:', total_general])
    total_row = sheet[sheet.max_row]
    total_row[4].font = Font(bold=True)
    total_row[5].font = Font(bold=True)
    total_row[5].fill = PatternFill(start_color="34a853", end_color="34a853", fill_type="solid")
    total_row[5].font = Font(bold=True, color="FFFFFF")
    
    # Statistics
    sheet.append([])
    stats_row = sheet.max_row + 1
    sheet.cell(row=stats_row, column=1).value = 'STATISTICS'
    sheet.cell(row=stats_row, column=1).font = Font(bold=True)
    stats_row += 1
    sheet.cell(row=stats_row, column=1).value = 'Total Sales:'
    sheet.cell(row=stats_row, column=2).value = len(sales)
    stats_row += 1
    sheet.cell(row=stats_row, column=1).value = 'Total Revenue:'
    sheet.cell(row=stats_row, column=2).value = f"$ {total_general:.2f}"
    stats_row += 1
    if len(sales) > 0:
        sheet.cell(row=stats_row, column=1).value = 'Average Ticket:'
        sheet.cell(row=stats_row, column=2).value = f"$ {total_general / len(sales):.2f}"
    
    # Adjust column width
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    
    # Save file
    wb.save(filepath)
    return filepath

